import pandas as pd
import requests
import time
from openpyxl import load_workbook
from tqdm import tqdm

# Configuración
client_id = ""
client_secret = ""
region = 'us'
locale = 'es_MX'

# Obtener token de acceso
def obtener_token(client_id, client_secret):
    url = f'https://{region}.battle.net/oauth/token'
    data = {'grant_type': 'client_credentials'}
    response = requests.post(url, data=data, auth=(client_id, client_secret))
    response.raise_for_status()
    return response.json()['access_token']

# Obtener datos del item
def obtener_datos_item(token, item_id):
    url = f'https://{region}.api.blizzard.com/data/wow/item/{item_id}'
    headers = {'Authorization': f'Bearer {token}'}
    params = {'namespace': f'static-{region}', 'locale': locale}
    response = requests.get(url, headers=headers, params=params)
    response.raise_for_status()
    return response.json()

# Actualizar el archivo Excel con las nuevas columnas
def actualizar_excel(file_path, item_id, category, subcategory):
    book = load_workbook(file_path)
    sheet = book.active
    
    # Find the row with the item_id
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
        if row[0].value == item_id:
            sheet.cell(row=row[0].row, column=3, value=category)
            sheet.cell(row=row[0].row, column=4, value=subcategory)
            break
    
    book.save(file_path)

def main():
    # Load the Excel file
    file_path = 'item_cache.xlsx'
    df = pd.read_excel(file_path)

    # Add columns for category and subcategory if they don't exist
    if 'category' not in df.columns:
        df['category'] = ''
    if 'subcategory' not in df.columns:
        df['subcategory'] = ''
    df.to_excel(file_path, index=False)

    # Get the token
    token = obtener_token(client_id, client_secret)

    for index, row in tqdm(df.iterrows(), total=len(df), desc="Processing items"):
        item_id = row['item_id']
        # Skip if category and subcategory are already filled
        if row['category'] != 0 and row['subcategory'] != 0:
            continue
        
        try:
            # Get item data from the API
            item_data = obtener_datos_item(token, item_id)
            category = item_data['item_class']['name'] if 'item_class' in item_data else 'Unknown'
            subcategory = item_data['item_subclass']['name'] if 'item_subclass' in item_data else 'Unknown'
            
            # Update Excel file with the category and subcategory
            df.at[index, 'category'] = category
            df.at[index, 'subcategory'] = subcategory
            df.to_excel(file_path, index=False)
            
        except Exception as e:
            print(f"Error updating item_id {item_id}: {e}")
        
        # Sleep to avoid hitting API rate limits
        time.sleep(0.05)

if __name__ == '__main__':
    main()